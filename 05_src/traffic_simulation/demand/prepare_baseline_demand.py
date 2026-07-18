"""Build governed Ota Ward population and parcel-equivalent demand meshes."""

from __future__ import annotations

import argparse
import hashlib
import io
import json
import os
import re
import zipfile
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, ROUND_FLOOR, ROUND_HALF_UP, getcontext
from pathlib import Path
from typing import Any, Final, Mapping, Sequence
from zoneinfo import ZoneInfo

import geopandas as gpd
import pandas as pd
import yaml
from openpyxl import load_workbook
from pyproj import CRS
from shapely.geometry import Polygon

from traffic_simulation.network.study_areas import (
    REGISTRY_FIELDS,
    StudyArea,
    load_study_area,
    lookup_source,
)
from traffic_simulation.paths import REPOSITORY_ROOT, SOURCE_REGISTRY


getcontext().prec = 40

SCHEMA_VERSION: Final = 1
CONFIG_PATH: Final = (
    REPOSITORY_ROOT
    / "reproducibility"
    / "config"
    / "traffic_simulation"
    / "baseline_demand.yml"
)
JST: Final = ZoneInfo("Asia/Tokyo")
MESH_CODE_PATTERN: Final = re.compile(r"^[0-9]{9}$")
SHA256_PATTERN: Final = re.compile(r"^[0-9a-f]{64}$")
PROCESSING_SCRIPT: Final = (
    "05_src/traffic_simulation/demand/prepare_baseline_demand.py"
)
ROOT_FIELDS: Final = frozenset({"schema_version", "baseline_demand"})
BASELINE_FIELDS: Final = frozenset(
    {"version", "status", "region_id", "target_days", "population", "demand_proxy", "outputs"}
)
POPULATION_FIELDS: Final = frozenset(
    {
        "mesh_source_registry_id",
        "mesh_definition_registry_id",
        "mesh_crs",
        "mesh_member",
        "source_encoding",
        "mesh_code_column",
        "population_column",
        "ota_total_source_registry_id",
        "ota_total_sheet",
        "ota_total_cell",
        "expected_ota_total",
        "boundary_allocation",
        "integer_allocation",
        "tie_breaker",
    }
)
DEMAND_FIELDS: Final = frozenset(
    {
        "parcel_source_registry_id",
        "national_population_source_registry_id",
        "national_population_sheet",
        "national_population_cell",
        "national_population_unit_multiplier",
        "expected_national_population",
        "annual_parcel_count",
        "annual_days",
        "unit",
        "integer_allocation",
        "tie_breaker",
    }
)
OUTPUT_FIELDS: Final = frozenset({"population_and_demand", "quality_summary"})


class _UniqueKeyLoader(yaml.SafeLoader):
    """YAML loader that rejects duplicate mapping keys."""


def _construct_unique_mapping(
    loader: _UniqueKeyLoader, node: yaml.nodes.MappingNode, deep: bool = False
) -> dict[Any, Any]:
    mapping: dict[Any, Any] = {}
    for key_node, value_node in node.value:
        key = loader.construct_object(key_node, deep=deep)
        if key in mapping:
            raise ValueError(f"duplicate YAML key: {key}")
        mapping[key] = loader.construct_object(value_node, deep=deep)
    return mapping


_UniqueKeyLoader.add_constructor(
    yaml.resolver.BaseResolver.DEFAULT_MAPPING_TAG,
    _construct_unique_mapping,
)


@dataclass(frozen=True, slots=True)
class BaselineDemandConfig:
    """Strict, immutable inputs for one population-demand preparation run."""

    version: int
    region_id: str
    target_days: int
    population: Mapping[str, Any]
    demand_proxy: Mapping[str, Any]
    outputs: Mapping[str, str]
    config_path: Path
    config_sha256: str


def _mapping(value: Any, label: str) -> dict[str, Any]:
    if not isinstance(value, dict) or not all(isinstance(key, str) for key in value):
        raise ValueError(f"{label} must be a string-keyed mapping")
    return value


def _exact_fields(value: Mapping[str, Any], expected: frozenset[str], label: str) -> None:
    missing = expected - set(value)
    unknown = set(value) - expected
    if missing:
        raise ValueError(f"{label} lacks fields: {sorted(missing)}")
    if unknown:
        raise ValueError(f"{label} has unknown fields: {sorted(unknown)}")


def _text(value: Any, label: str) -> str:
    if not isinstance(value, str) or not value.strip():
        raise ValueError(f"{label} must be a non-empty string")
    return value


def _positive_integer(value: Any, label: str) -> int:
    if isinstance(value, bool) or not isinstance(value, int) or value <= 0:
        raise ValueError(f"{label} must be a positive integer")
    return value


def _relative_output(value: Any, label: str) -> str:
    text = _text(value, label)
    path = Path(text)
    if path.is_absolute() or ".." in path.parts:
        raise ValueError(f"{label} must be a safe repository-relative path")
    if not text.startswith("03_data/processed/traffic_simulation/"):
        raise ValueError(f"{label} must be under the traffic-simulation processed root")
    return text


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def load_config(path: Path = CONFIG_PATH) -> BaselineDemandConfig:
    """Load the versioned configuration with no implicit fallback values."""

    raw_bytes = path.read_bytes()
    try:
        document = yaml.load(raw_bytes.decode("utf-8"), Loader=_UniqueKeyLoader)
    except (UnicodeDecodeError, yaml.YAMLError) as exc:
        raise ValueError(f"invalid UTF-8 YAML in {path}") from exc
    root = _mapping(document, "root")
    _exact_fields(root, ROOT_FIELDS, "root")
    if root["schema_version"] != SCHEMA_VERSION or isinstance(root["schema_version"], bool):
        raise ValueError(f"unsupported schema_version: {root['schema_version']!r}")

    baseline = _mapping(root["baseline_demand"], "baseline_demand")
    _exact_fields(baseline, BASELINE_FIELDS, "baseline_demand")
    version = _positive_integer(baseline["version"], "version")
    if baseline["status"] != "active":
        raise ValueError("baseline_demand.status must be active")
    region_id = _text(baseline["region_id"], "region_id")
    target_days = _positive_integer(baseline["target_days"], "target_days")

    population = _mapping(baseline["population"], "population")
    demand = _mapping(baseline["demand_proxy"], "demand_proxy")
    outputs = _mapping(baseline["outputs"], "outputs")
    _exact_fields(population, POPULATION_FIELDS, "population")
    _exact_fields(demand, DEMAND_FIELDS, "demand_proxy")
    _exact_fields(outputs, OUTPUT_FIELDS, "outputs")

    for key in (
        "mesh_source_registry_id",
        "mesh_definition_registry_id",
        "mesh_member",
        "source_encoding",
        "mesh_code_column",
        "population_column",
        "ota_total_source_registry_id",
        "ota_total_sheet",
        "ota_total_cell",
    ):
        _text(population[key], f"population.{key}")
    if CRS.from_user_input(_text(population["mesh_crs"], "population.mesh_crs")).to_epsg() != 6668:
        raise ValueError("population.mesh_crs must be EPSG:6668 for JGD2011 mesh codes")
    _positive_integer(population["expected_ota_total"], "expected_ota_total")
    if population["boundary_allocation"] != "area_weighted":
        raise ValueError("only area_weighted boundary allocation is supported")
    if population["integer_allocation"] != "largest_remainder":
        raise ValueError("population integer allocation must be largest_remainder")
    if population["tie_breaker"] != "mesh_code_ascending":
        raise ValueError("population tie breaker must be mesh_code_ascending")

    for key in (
        "parcel_source_registry_id",
        "national_population_source_registry_id",
        "national_population_sheet",
        "national_population_cell",
        "unit",
    ):
        _text(demand[key], f"demand_proxy.{key}")
    for key in (
        "national_population_unit_multiplier",
        "expected_national_population",
        "annual_parcel_count",
        "annual_days",
    ):
        _positive_integer(demand[key], f"demand_proxy.{key}")
    if demand["unit"] != "parcel_equivalent":
        raise ValueError("demand_proxy.unit must be parcel_equivalent")
    if demand["integer_allocation"] != "largest_remainder":
        raise ValueError("demand integer allocation must be largest_remainder")
    if demand["tie_breaker"] != "mesh_code_ascending":
        raise ValueError("demand tie breaker must be mesh_code_ascending")

    validated_outputs = {
        key: _relative_output(value, f"outputs.{key}") for key, value in outputs.items()
    }
    return BaselineDemandConfig(
        version=version,
        region_id=region_id,
        target_days=target_days,
        population=dict(population),
        demand_proxy=dict(demand),
        outputs=validated_outputs,
        config_path=path,
        config_sha256=hashlib.sha256(raw_bytes).hexdigest(),
    )


def _source_path(row: Mapping[str, str]) -> Path:
    relative = Path(row["local_raw_path"])
    if relative.is_absolute() or ".." in relative.parts:
        raise ValueError("source registry contains an unsafe raw path")
    path = (REPOSITORY_ROOT / relative).resolve()
    try:
        path.relative_to(REPOSITORY_ROOT)
    except ValueError as exc:
        raise ValueError("source registry path escapes repository") from exc
    return path


def verify_source(source_id: str, registry_path: Path = SOURCE_REGISTRY) -> tuple[Path, str]:
    """Resolve one registered source and enforce its registered SHA-256.

    ``original_filename`` records the distributor-side name and may differ
    from the basename chosen for ``local_raw_path``.
    """

    row = lookup_source(source_id, registry_path)
    path = _source_path(row)
    if not path.is_file():
        raise FileNotFoundError(f"registered source does not exist: {path}")
    if not SHA256_PATTERN.fullmatch(row["sha256"]):
        raise ValueError(f"invalid registered SHA-256 for {source_id}")
    actual = sha256_file(path)
    if actual != row["sha256"]:
        raise ValueError(f"raw SHA-256 hash mismatch for {source_id}: {actual} != {row['sha256']}")
    return path, actual


def read_population_mesh(
    path: Path,
    *,
    member: str,
    encoding: str,
    mesh_code_column: str,
    population_column: str,
) -> tuple[pd.DataFrame, list[str]]:
    """Read only the configured official CSV member from the immutable ZIP."""

    if not zipfile.is_zipfile(path):
        raise ValueError(f"population mesh source is not a ZIP: {path}")
    with zipfile.ZipFile(path) as archive:
        names = archive.namelist()
        if names.count(member) != 1:
            raise ValueError(f"expected one configured mesh member {member!r}, found {names.count(member)}")
        if Path(member).is_absolute() or ".." in Path(member).parts:
            raise ValueError("unsafe configured mesh member")
        payload = archive.read(member)
    frame = pd.read_csv(
        io.BytesIO(payload),
        encoding=encoding,
        skiprows=[1],
        dtype={mesh_code_column: "string", population_column: "string"},
        low_memory=False,
    )
    missing = {mesh_code_column, population_column} - set(frame.columns)
    if missing:
        raise ValueError(f"population mesh lacks columns: {sorted(missing)}")
    codes = frame[mesh_code_column].astype("string")
    if codes.isna().any() or not codes.map(lambda value: bool(MESH_CODE_PATTERN.fullmatch(str(value)))).all():
        raise ValueError("population mesh contains an invalid 500m mesh code")
    if codes.duplicated().any():
        raise ValueError("population mesh contains duplicate mesh codes")
    population = pd.to_numeric(frame[population_column], errors="coerce")
    if population.isna().any() or (population < 0).any() or not (population % 1 == 0).all():
        raise ValueError("population total column contains a missing or invalid value")
    result = pd.DataFrame(
        {
            "mesh_code": codes,
            "census_population_2020": population.astype("int64"),
        }
    )
    return result, names


def mesh_500m_polygon(mesh_code: str) -> Polygon:
    """Decode one standard 9-digit half-mesh code into its JGD2011 cell."""

    if not MESH_CODE_PATTERN.fullmatch(mesh_code):
        raise ValueError(f"invalid 500m mesh code: {mesh_code!r}")
    latitude_degree = int(mesh_code[0:2])
    longitude_degree = int(mesh_code[2:4])
    latitude_1km = int(mesh_code[4])
    longitude_1km = int(mesh_code[5])
    latitude_100m_digit = int(mesh_code[6])
    longitude_100m_digit = int(mesh_code[7])
    quadrant = int(mesh_code[8])
    if latitude_1km > 7 or longitude_1km > 7:
        raise ValueError(f"invalid 1km subdivision in mesh code: {mesh_code}")
    if latitude_100m_digit > 9 or longitude_100m_digit > 9 or quadrant not in {1, 2, 3, 4}:
        raise ValueError(f"invalid subdivision in mesh code: {mesh_code}")

    south = (
        Decimal(latitude_degree) * Decimal(2) / Decimal(3)
        + Decimal(latitude_1km) / Decimal(12)
        + Decimal(latitude_100m_digit) / Decimal(120)
    )
    west = (
        Decimal(100 + longitude_degree)
        + Decimal(longitude_1km) / Decimal(8)
        + Decimal(longitude_100m_digit) / Decimal(80)
    )
    half_height = Decimal(1) / Decimal(240)
    half_width = Decimal(1) / Decimal(160)
    if quadrant in {3, 4}:
        south += half_height
    if quadrant in {2, 4}:
        west += half_width
    north = south + half_height
    east = west + half_width
    return Polygon(
        [
            (float(west), float(south)),
            (float(east), float(south)),
            (float(east), float(north)),
            (float(west), float(north)),
            (float(west), float(south)),
        ]
    )


def largest_remainder(
    expected: Sequence[Decimal], *, total: int, keys: Sequence[str]
) -> list[int]:
    """Allocate a fixed integer total by floor, remainder, then key order."""

    if len(expected) != len(keys):
        raise ValueError("expected values and keys must have equal length")
    if len(keys) != len(set(keys)):
        raise ValueError("largest-remainder keys must be unique")
    if total < 0 or any(value < 0 for value in expected):
        raise ValueError("largest-remainder inputs must be non-negative")
    floors = [int(value.to_integral_value(rounding=ROUND_FLOOR)) for value in expected]
    remaining = total - sum(floors)
    if remaining < 0 or remaining > len(expected):
        raise ValueError("requested total is incompatible with expected values")
    order = sorted(
        range(len(expected)),
        key=lambda index: (-(expected[index] - Decimal(floors[index])), keys[index]),
    )
    result = floors[:]
    for index in order[:remaining]:
        result[index] += 1
    if sum(result) != total:
        raise AssertionError("largest-remainder allocation did not preserve total")
    return result


def _read_xlsx_number(path: Path, *, sheet: str, cell: str, multiplier: int = 1) -> int:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet not in workbook.sheetnames:
            raise ValueError(f"XLSX lacks configured sheet {sheet!r}")
        value = workbook[sheet][cell].value
    finally:
        workbook.close()
    if isinstance(value, bool) or not isinstance(value, (int, float)) or not float(value).is_integer():
        raise ValueError(f"XLSX cell {sheet}!{cell} is not an integer")
    return int(value) * multiplier


def build_population_and_demand(
    mesh: pd.DataFrame,
    *,
    area: StudyArea,
    target_population: int,
    q_base: Decimal,
    target_days: int,
    mesh_crs: str,
) -> tuple[gpd.GeoDataFrame, dict[str, Any]]:
    """Clip population meshes, rescale population, and allocate demand."""

    geometries = [mesh_500m_polygon(code) for code in mesh["mesh_code"]]
    cells = gpd.GeoDataFrame(mesh.copy(), geometry=geometries, crs=mesh_crs)
    west, south, east, north = area.acquisition_bbox
    cells = cells.cx[west:east, south:north].copy()
    if cells.empty:
        raise ValueError("population mesh has no cells in the study-area bbox")
    api_cells = cells.to_crs(area.api_crs)
    metric = cells.to_crs(area.metric_crs)
    mesh_area = metric.geometry.area
    intersections = metric.geometry.intersection(area.metric_boundary)
    overlap_area = intersections.area
    positive = overlap_area > 0
    metric = metric.loc[positive].copy()
    intersections = intersections.loc[positive]
    mesh_area = mesh_area.loc[positive]
    overlap_area = overlap_area.loc[positive]
    if metric.empty:
        raise ValueError("population mesh has no positive-area intersection with study area")

    # Classify containment in the geographic CRS.  Reprojecting a long shared
    # edge as one segment versus several shorter segments introduces a small
    # artificial area difference, so a metric-area tolerance is not a reliable
    # test for full containment.
    full = api_cells.loc[positive].geometry.map(area.api_boundary.covers)
    intersections.loc[full] = metric.loc[full].geometry
    overlap_area.loc[full] = mesh_area.loc[full]
    overlap_ratio = (overlap_area / mesh_area).clip(lower=0.0, upper=1.0)
    census_population = metric["census_population_2020"].astype("int64")
    weighted = census_population.astype("float64") * overlap_ratio
    if float(weighted.sum()) <= 0:
        raise ValueError("area-weighted census population is not positive")

    expected_population = [
        Decimal(target_population) * Decimal(str(value)) / Decimal(str(weighted.sum()))
        for value in weighted
    ]
    keys = [str(value) for value in metric["mesh_code"]]
    allocated_population = largest_remainder(
        expected_population, total=target_population, keys=keys
    )
    demand_expected = [Decimal(value) * q_base * Decimal(target_days) for value in allocated_population]
    demand_total_expected = Decimal(target_population) * q_base * Decimal(target_days)
    demand_total = int(demand_total_expected.quantize(Decimal("1"), rounding=ROUND_HALF_UP))
    allocated_demand = largest_remainder(demand_expected, total=demand_total, keys=keys)

    centroid_included = metric.geometry.centroid.map(area.metric_boundary.covers)
    centroid_population = int(census_population.loc[centroid_included].sum())
    area_weighted_population = float(weighted.sum())

    output = gpd.GeoDataFrame(
        {
            "region_id": area.region_id,
            "mesh_code": keys,
            "census_population_2020": census_population.to_numpy(),
            "mesh_area_m2": mesh_area.to_numpy(),
            "boundary_overlap_area_m2": overlap_area.to_numpy(),
            "boundary_overlap_ratio": overlap_ratio.to_numpy(),
            "boundary_class": ["full" if value else "partial" for value in full],
            "area_weighted_population_2020": weighted.to_numpy(),
            "population_2024_expected": [float(value) for value in expected_population],
            "population_2024": allocated_population,
            "demand_expected_parcel_equivalent": [float(value) for value in demand_expected],
            "demand_parcel_equivalent": allocated_demand,
            "geometry": intersections.to_numpy(),
        },
        geometry="geometry",
        crs=area.metric_crs,
    ).sort_values("mesh_code", ignore_index=True)

    summary = {
        "region_id": area.region_id,
        "target_days": target_days,
        "intersecting_mesh_count": int(len(output)),
        "full_mesh_count": int((output["boundary_class"] == "full").sum()),
        "partial_boundary_mesh_count": int((output["boundary_class"] == "partial").sum()),
        "minimum_overlap_ratio": float(output["boundary_overlap_ratio"].min()),
        "maximum_overlap_ratio": float(output["boundary_overlap_ratio"].max()),
        "area_weighted_census_population_2020": area_weighted_population,
        "centroid_inclusion_census_population_2020": centroid_population,
        "boundary_method_population_difference": area_weighted_population - centroid_population,
        "target_population_2024": target_population,
        "allocated_population_2024": int(output["population_2024"].sum()),
        "q_base": str(q_base),
        "expected_demand_parcel_equivalent": str(demand_total_expected),
        "allocated_demand_parcel_equivalent": int(output["demand_parcel_equivalent"].sum()),
        "population_positive_mesh_count": int((output["population_2024"] > 0).sum()),
        "demand_positive_mesh_count": int((output["demand_parcel_equivalent"] > 0).sum()),
        "output_crs": output.crs.to_string() if output.crs else None,
    }
    return output, summary


def output_paths(config: BaselineDemandConfig) -> tuple[Path, Path]:
    population_and_demand = (
        REPOSITORY_ROOT / config.outputs["population_and_demand"]
    ).resolve()
    quality_summary = (REPOSITORY_ROOT / config.outputs["quality_summary"]).resolve()
    return population_and_demand, quality_summary


def write_outputs(
    frame: gpd.GeoDataFrame,
    summary: Mapping[str, Any],
    *,
    parquet_path: Path,
    summary_path: Path,
) -> dict[str, Any]:
    """Write both products atomically and refuse to overwrite either one."""

    existing = [path for path in (parquet_path, summary_path) if path.exists()]
    if existing:
        raise FileExistsError(f"refusing to overwrite output: {existing[0]}")
    parquet_path.parent.mkdir(parents=True, exist_ok=True)
    summary_path.parent.mkdir(parents=True, exist_ok=True)
    parquet_part = parquet_path.with_name(parquet_path.name + ".part")
    summary_part = summary_path.with_name(summary_path.name + ".part")
    for path in (parquet_part, summary_part):
        if path.exists():
            path.unlink()
    parquet_committed = False
    summary_committed = False
    try:
        frame.to_parquet(parquet_part, index=False)
        final_summary = dict(summary)
        final_summary["population_and_demand_sha256"] = sha256_file(parquet_part)
        summary_part.write_text(
            json.dumps(final_summary, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
            encoding="utf-8",
        )
        os.replace(parquet_part, parquet_path)
        parquet_committed = True
        os.replace(summary_part, summary_path)
        summary_committed = True
        return final_summary
    finally:
        for path in (parquet_part, summary_part):
            if path.exists():
                path.unlink()
        if not summary_committed:
            if parquet_committed and parquet_path.exists():
                parquet_path.unlink()
            if summary_path.exists():
                summary_path.unlink()


def run(
    *,
    config_path: Path = CONFIG_PATH,
    registry_path: Path = SOURCE_REGISTRY,
) -> tuple[Path, Path, dict[str, Any]]:
    """Verify all governed sources, calculate meshes, and write products."""

    config = load_config(config_path)
    population = config.population
    demand = config.demand_proxy
    source_ids = (
        population["mesh_source_registry_id"],
        population["mesh_definition_registry_id"],
        population["ota_total_source_registry_id"],
        demand["national_population_source_registry_id"],
        demand["parcel_source_registry_id"],
    )
    sources = {source_id: verify_source(source_id, registry_path) for source_id in source_ids}

    mesh_path = sources[population["mesh_source_registry_id"]][0]
    mesh, members = read_population_mesh(
        mesh_path,
        member=population["mesh_member"],
        encoding=population["source_encoding"],
        mesh_code_column=population["mesh_code_column"],
        population_column=population["population_column"],
    )
    ota_path = sources[population["ota_total_source_registry_id"]][0]
    target_population = _read_xlsx_number(
        ota_path,
        sheet=population["ota_total_sheet"],
        cell=population["ota_total_cell"],
    )
    if target_population != population["expected_ota_total"]:
        raise ValueError("Ota total population does not match configured expected value")
    national_path = sources[demand["national_population_source_registry_id"]][0]
    national_population = _read_xlsx_number(
        national_path,
        sheet=demand["national_population_sheet"],
        cell=demand["national_population_cell"],
        multiplier=demand["national_population_unit_multiplier"],
    )
    if national_population != demand["expected_national_population"]:
        raise ValueError("national population does not match configured expected value")
    q_base = (
        Decimal(demand["annual_parcel_count"])
        / Decimal(national_population)
        / Decimal(demand["annual_days"])
    )
    area = load_study_area(
        config.region_id,
        registry_path=registry_path,
    )
    frame, summary = build_population_and_demand(
        mesh,
        area=area,
        target_population=target_population,
        q_base=q_base,
        target_days=config.target_days,
        mesh_crs=population["mesh_crs"],
    )
    summary.update(
        {
            "schema_version": SCHEMA_VERSION,
            "baseline_demand_version": config.version,
            "config_sha256": config.config_sha256,
            "source_sha256": {source_id: digest for source_id, (_, digest) in sources.items()},
            "mesh_archive_member": population["mesh_member"],
            "mesh_archive_other_members": [name for name in members if name != population["mesh_member"]],
            "generated_at": datetime.now(JST).isoformat(timespec="seconds"),
        }
    )
    parquet_path, summary_path = output_paths(config)
    final_summary = write_outputs(
        frame,
        summary,
        parquet_path=parquet_path,
        summary_path=summary_path,
    )
    return parquet_path, summary_path, final_summary


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Build Ota Ward population and parcel-equivalent demand meshes."
    )
    parser.add_argument("--config", type=Path, default=CONFIG_PATH)
    return parser


def main(argv: Sequence[str] | None = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)
    try:
        parquet_path, summary_path, summary = run(config_path=args.config)
    except (OSError, ValueError, yaml.YAMLError) as exc:
        parser.exit(1, f"error: {exc}\n")
    print(f"intersecting meshes: {summary['intersecting_mesh_count']}")
    print(f"partial boundary meshes: {summary['partial_boundary_mesh_count']}")
    print(f"population: {summary['allocated_population_2024']}")
    print(f"parcel-equivalent demand: {summary['allocated_demand_parcel_equivalent']}")
    print(f"population and demand: {parquet_path.relative_to(REPOSITORY_ROOT)}")
    print(f"quality summary: {summary_path.relative_to(REPOSITORY_ROOT)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
